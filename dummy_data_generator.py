#  Generate an excel file with data in following manner
# Columns -

# Question id
# Taxanomy (Category)
# Type of Queston
# Question
# Option 1
# Option 2
# Option 3
# Option 4
# Correct Option
# Marks

from openpyxl import *

workbook = Workbook()
sheet = workbook.active

sheet.cell(row=1, column=1).value = "Question id"
sheet.cell(row=1, column=2).value = "Taxanomy (Category)"
sheet.cell(row=1, column=3).value = "Type of Queston"
sheet.cell(row=1, column=4).value = "Question"
sheet.cell(row=1, column=5).value = "Option 1"
sheet.cell(row=1, column=6).value = "Option 2"
sheet.cell(row=1, column=7).value = "Option 3"
sheet.cell(row=1, column=8).value = "Option 4"
sheet.cell(row=1, column=9).value = "Correct Option"
sheet.cell(row=1, column=10).value = "Marks"

##################################################################
########################## Parameters ############################
##################################################################

subjects = ['Drill', 'WT', 'PD', 'LD', 'DM', 'SSCD', 'HH', 'EAC', 'OT', 'GA']
categories = ['Und', 'Hot', 'Rem', 'App', 'Emd']
types = ['VSA', 'SA', 'LA1', 'LA2', 'ET']
marks_of_each_type = [1, 2, 3, 4, 6]
no_of_dummy_questions_for_each_type = 10

##################################################################
###################### Generate dummy data #######################
##################################################################

for subject in subjects:
    count = 1
    for category in categories:
        for index, type in enumerate(types):
            for i in range(1, no_of_dummy_questions_for_each_type + 1):
                sheet.append([
                    f"{subject}.{count}",
                    category,
                    type,
                    f"{subject}_{category}_{type}_Q{i}",
                    "NULL",
                    "NULL",
                    "NULL",
                    "NULL",
                    "NULL",
                    marks_of_each_type[index]
                ])
                count += 1

workbook.save("dummy_data.xlsx")
