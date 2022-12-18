import random
from openpyxl import load_workbook
from fpdf import FPDF, HTMLMixin
from pdfPlugin import create_table

# LOAD EXCEL
workbook = load_workbook(filename='dummy_data.xlsx', data_only=True)
sheet = workbook.active

# LOAD DATA
questions = {}
for question in sheet.iter_rows(1, values_only=True, max_row=2501):
    # print(question)
    subject = question[0].split('.')[0]
    category = question[1]
    type = question[2]
    question_text = question[3]
    question_text_hindi = question[4]
    marks = question[10]
    if subject not in questions:
        questions[subject] = {}
    if category not in questions[subject]:
        questions[subject][category] = {}
    if type not in questions[subject][category]:
        questions[subject][category][type] = []
    questions[subject][category][type].append(
        [question_text, question_text_hindi, marks])

# GIVEN PAPER CONSTRAINTS
constraints = {
    'Drill': {
        'Und': {
            'VSA': 5,
            'LA1': 1,
            'ET': 3
        },
        'Hot': {
            'VSA': 3,
            'LA1': 2,
        },
        'Rem': {
            'SA': 2,
        },
        'App': {
            'SA': 3,
            'LA1': 1,
        },
        'Emd': {
            'LA2': 3,
        }
    },
    'WT': {
        'Und': {
            'VSA': 5,
        },
        'Hot': {
            'VSA': 5,
            'LA1': 2,
            'LA2': 2,
            'ET': 3
        },
        'App': {
            'SA': 5,
            'ET': 1
        },
        'Emd': {
            'VSA': 1,
            'LA1': 2,
        }
    },
    'PD': {
        'Und': {
            'SA': 5,
            'LA1': 2,
        },
        'Hot': {
            'SA': 5,
            'ET': 1
        },
        'App': {
            'LA1': 2,
            'LA2': 3,
            'ET': 2
        },
        'Emd': {
            'VSA': 7,
            'LA1': 2,
        }
    },
    'LD': {
        'Hot': {
            'LA1': 1,
        },
        'Rem': {
            'VSA': 5,
        },
        'App': {
            'SA': 1,
        },
    },
    'DM': {
        'Und': {
            'ET': 1
        },
        'Hot': {
            'LA1': 1,
        },
        'App': {
            'LA2': 1,
        },
        'Emd': {
            'SA': 3,
        }
    },
    'SSCD': {
        'Und': {
            'SA': 5,
        },
        'Hot': {
            'LA1': 4,
        },
        'App': {
            'VSA': 5,
        },
        'Emd': {
            'LA2': 2,
        }
    },
    'HH': {
        'Und': {
            'LA1': 2,
        },
        'Hot': {
            'LA2': 2,
            'ET': 1
        },
        'Rem': {
            'VSA': 1,
            'SA': 2,
        },
        'App': {
            'SA': 5,
        },
    },
    'EAC': {
        'Hot': {
            'VSA': 2,
            'LA1': 1,
        },
        'Rem': {
            'VSA': 5,
            'SA': 1,
        },
        'Emd': {
            'VSA': 2,
            'SA': 2,
            'LA1': 1,
        }
    },
    'OT': {
        'Und': {
            'ET': 1
        },
        'App': {
            'SA': 1,
            'LA1': 1,
            'LA2': 1,
        }
    },
    'GA': {
        'App': {
            'LA1': 1,
            'ET': 2
        },
    }
}

# GENERATED PAPER EXPORT FORMAT
exportFormat = "pdf"  # pdf or txt or console

# LOGIC


class PDF(FPDF, HTMLMixin):
    pass


total_marks = 0
pdf = PDF(orientation='P', unit='mm')
pdf.add_font("akshar", fname="./Akshar-Unicode.ttf")
# line_height = pdf.font_size * 2.5
# col_width = pdf.epw / 4  # distribute content evenly
cell_widths = "uneven"
col_width = pdf.epw * 0.9

if exportFormat == "pdf":
    pdf.add_page()
    pdf.set_font('helvetica', 'B', 12)
    pdf.cell(w=0, txt='NCC C Certificate Exam', align='C')

metadata = [["Topology", "VSA", "SA", "LA1",
             "LA2", "ET", "Total Questions", "Total Marks"]]
metaGlobalObject = {
    "Rem": [0, 0, 0, 0, 0],
    "Und": [0, 0, 0, 0, 0],
    "App": [0, 0, 0, 0, 0],
    "Hot": [0, 0, 0, 0, 0],
    "Emd": [0, 0, 0, 0, 0]
}
typeCast = {
    "VSA": 0,
    "SA": 1,
    "LA1": 2,
    "LA2": 3,
    "ET": 4
}
metaGlobalSubjectObject = {key: 0 for key in constraints.keys()}


def calcMarks(arr: list):
    if len(arr) != 5:
        return 0
    return arr[0] * 1 + arr[1] * 2 + arr[2] * 3 + arr[3] * 4 + arr[4] * 6


for subject in constraints:

    if exportFormat == "console":
        print(subject)
        print('-' * len(subject))
    elif exportFormat == "txt":
        paperData = ''
        question_paper = open("question_paper_" + subject + ".txt", "w")
        paperData += subject + '\n' + ('-' * len(subject)) + '\n'
    #     pdf.set_font('helvetica', 'B', 16)
    #     pdf.cell(w=0, txt=subject, align='C')
    #     pdf.ln(15)

    subject_marks = 0
    question_count = 1
    table_data = [["S.No.", "Question", "Marks"]]
    # language_data = []
    for category in constraints[subject]:
        if(exportFormat == "console"):
            print(category)
        for type in constraints[subject][category]:
            no_of_questions = constraints[subject][category][type]
            metaGlobalObject[category][typeCast[type]] += no_of_questions
            selected_questions = random.sample(
                questions[subject][category][type], no_of_questions)
            for question in selected_questions:

                if(exportFormat == "console"):
                    print(
                        f"Question {question_count}. {question[0]}", '\t\t', question[2])
                    print(
                        f"Question {question_count}. {question[1]}", '\t\t', question[2])
                elif(exportFormat == "txt"):
                    paperData += f"Question {question_count}. {question[0]}\t\t{question[2]} Marks\n"
                    paperData += f"Question {question_count}. {question[1]}\t\t{question[2]} Marks\n"
                elif(exportFormat == "pdf"):
                    table_data.append(
                        [str(question_count), str(question[0]), str(question[2])])
                    table_data.append(
                        [str(question_count), str(question[1]), str(question[2])])
                    # language_data.append(question[0])
                question_count += 1
                subject_marks += question[2]
    total_marks += subject_marks
    metaGlobalSubjectObject[subject] += subject_marks
    if exportFormat == "console":
        print()
        print(f"Total marks for {subject}: {subject_marks}")
        print('-----------------------------------------------------')
        print()
    elif exportFormat == "txt":
        paperData += f"\nTotal marks for {subject}: {subject_marks}\n"
        paperData += '-----------------------------------------------------\n'
        paperData += '-----------------------------------------------------\n'
        question_paper.write(paperData)
        question_paper.close()
    elif exportFormat == "pdf":
        create_table(pdf, table_data, subject, f"Total marks for {subject}: {subject_marks}", 12, 16, 'L', 'L', [
                     13, 160, 15], 'x_default')
        pdf.add_page()

if exportFormat == "console":
    print('-----------------------------------------------------')
    print(f"Total marks: {total_marks}")
    print('-----------------------------------------------------')
    print(metadata)

elif exportFormat == "pdf":
    pdf.set_font('helvetica', 'B', 12)
    pdf.cell(0, 5, '-----------------------------------------------------')
    pdf.ln()
    pdf.cell(0, 10, f"Total marks: {total_marks}")
    pdf.ln()
    pdf.cell(0, 5, '-----------------------------------------------------')
    pdf.ln()
    metadata.append(["Rem", *metaGlobalObject["Rem"],
                    sum(metaGlobalObject["Rem"]), calcMarks(metaGlobalObject["Rem"])])
    metadata.append(["Und", *metaGlobalObject["Und"],
                    sum(metaGlobalObject["Und"]), calcMarks(metaGlobalObject["Und"])])
    metadata.append(["App", *metaGlobalObject["App"],
                    sum(metaGlobalObject["App"]), calcMarks(metaGlobalObject["App"])])
    metadata.append(["Hot", *metaGlobalObject["Hot"],
                    sum(metaGlobalObject["Hot"]), calcMarks(metaGlobalObject["Hot"])])
    metadata.append(["Emd", *metaGlobalObject["Emd"],
                    sum(metaGlobalObject["Emd"]), calcMarks(metaGlobalObject["Emd"])])
    total_vsa = sum([x[1] for x in metadata[1:]])
    total_sa = sum([x[2] for x in metadata[1:]])
    total_la1 = sum([x[3] for x in metadata[1:]])
    total_la2 = sum([x[4] for x in metadata[1:]])
    total_et = sum([x[5] for x in metadata[1:]])
    metadata.append(["Total Questions",
                     total_vsa,
                     total_sa,
                     total_la1,
                     total_la2,
                     total_et,
                     total_vsa + total_sa + total_la1 + total_la2 + total_et,
                     total_marks])
    print(metadata)
    create_table(pdf, metadata, [], "Metadata", "", 12,
                 16, 'L', 'L', "uneven", 'x_default')

    print(metaGlobalSubjectObject)
    metaSubjectData = [["Subject", "Written"]]
    for subject in metaGlobalSubjectObject.keys():
        metaSubjectData.append([subject, metaGlobalSubjectObject[subject]])
    metaSubjectData.append(["Total", total_marks])
    create_table(pdf, metaSubjectData, [], "Subject Metadata", "", 12,
                 16, 'L', 'L', "uneven", 'x_default')
    pdf.output('question_paper.pdf', 'F')
